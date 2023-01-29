#!/usr/bin/env python3
#
# script genere_grilles_indiv.py
# v2023-01-22 🟢⚠❌📌‼❓🔷👉⌨️
# doc openpyxl : https://openpyxl.readthedocs.io


################################################################
# CONSTANTES :

# diplômes possibles :1
DIPLOMES        = {'31212': "Métiers de l'accueil",
                   '31213': "Met.com.ven.op.A Ani.ges.esp.com.",
                   '31214': "Met.com.ven.Op.B Pr.cl.va.of.com."}  # !!todo!! : manque 1 diplôme (CAP)
DIPLOMES_COURTS = {'31212': "bacpro_MA",
                   '31213': "bacpro_MVC_A_AGEC",
                   '31214': "bacpro_MVC_B_PC"}                    # !!todo!! : manque 1 diplôme (CAP)

# préfixe répertoire
INDIV_FOLDER    = 'candidats_'

# fichiers "source de données" : CYCLADE
CYCLADE_PREFIX = "cyclade"  # les fichiers CSV exportés de Cyclade doivent être commencés par ...

# old :
LIST_FILE       = 'DOSSIER ETABLISSEMENT CAP EPC.xlsx'
LIST_SHEET      = 'RECAPNOTES'
LIST_RANGE      = (12, 1, 46, 3)  # ie : début = L12, C1 ; fin = L46, C3


# fichiers "modèles" :
TEMPLATE_FILES  = ['31212.xlsx', '31213.xlsx', '31214.xlsx']
TEMPLATE_SHEET  = '1-Candidat, établissement'
TEMPLATE_DICT   = {'session': 'A3',
                   'etab': 'A4',
                   'UAI': 'A5',
                   
                   'nom': 'A7',
                   'prenom': 'A8',
                   'daten': 'A9',
                   'numcandidat': 'A10',
                   'division': 'A11',
                   'code': 'A12'}  # !!PB!! à modifier !
TEMPO           = 0.1
NSEP            = 32  # pour l'affichage des séparateurs
NEWLINE         = "\n"  # pour le saut de ligne dans le print de DOC

# pour sanitize()
CHAR_SEP        = "+"
CHAR_SUB        = "_"
MAJ             = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
SPE             = "ÀÁÂÃÄÅÆÇÈÉÊËÌÍÎÏÑÒÓÔÕÖŒÙÚÛÜ"
NBR             = "0123456789-"
LEG_CHAR        = MAJ + MAJ.lower() + SPE + SPE.lower() + NBR

DOC = f"""

* - * - * - * - * - * - * - * - * - * - * - * - * - * - * - *
  Script pour générer les fichiers individuels des candidats
* - * - * - * - * - * - * - * - * - * - * - * - * - * - * - *

Pour que ce script fonctionne correctement, il est nécessaire de vérifier
les prérequis suivants :
- Python 3 installé ; si ce message s'affiche, c'est sûrement le cas ;-).
- La bibliothèque Python 'openpyxl' est installée.
- Dans le même répertoire que le script, se trouvent :
  - 👉 les fichiers d'export Cyclade, qui possèdent l'extension ".csv"
    et dont le nom commence par "{CYCLADE_PREFIX}"
    Il peut y en avoir plusieurs (typiquement, un pour le CAP,
    un pour le bac pro) ; les candidats seront recherchés
    dans chacun d'entre eux ;
  - 👉 les fichiers modèles nécessaires :
{NEWLINE.join(f'        "{key}.xlsx" -> {value}' for key, value in DIPLOMES.items())}
    avec une feuille "{TEMPLATE_SHEET}".

Appuyez sur [Entrée] pour continuer, [CTRL+C] pour arrêter.

"""

################################################################

print(DOC)
input()

try:
    import sys
    import os
    import time
    import shutil
    import csv
except:
    print("❌ Une des bibliothèques standards est manquante !\n")
    sys.exit(1)

try:
    import openpyxl
except:
    print("❌ La bibliothèque suivante n'est pas installée : openpyxl\n")
    sys.exit(2)


def clear():
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')


def stamp():
    now = time.localtime()
    res = time.strftime("%Y%m%d_%H%M%S", now)
    return res


def sanitize(s):
    res = ""
    for letter in s:
        if letter in LEG_CHAR:
            res += letter
        else:
            res += CHAR_SUB
    return res


def touche():
    input("-" * NSEP + "\n⌨️ Appuyez sur la touche 'Entrée' pour continuer.\n" + "_" * NSEP + "\n")


def info(s):
    print(f"\n{'-'*NSEP}\nℹ️ : {s}\n{'_'*NSEP}\n")


################################################################
################################################################
################################################################


clear()
info("Début du traitement...")
# print("\n\n"+"-"*NSEP*2+"\nDébut du traitement...\n"+"-"*NSEP*2+"\n")


################################################################
# lire les données depuis le(s) fichier(s) Cyclade :


################################################################
# récupérer une liste de tous les fichiers du répertoire courant
files = os.listdir(".")
files = [f for f in files if os.path.isfile(f)]  # exclure les répertoires


################################################################
# construire une liste de tous les fichiers dont le nom commence par CYCLADE_PREFIX et d'extension .csv
files_cyclade = []
for f in files:
    if f[0:7].lower() == CYCLADE_PREFIX.lower() and f.split('.')[-1].lower() == 'csv':
        files_cyclade.append(f)


################################################################
# si aucun fichier cycladeXYZ.csv : sortie en erreur
if len(files_cyclade) == 0:
    print(f"❌ Fichier(s) Cyclade ({CYCLADE_PREFIX}XYZ.csv) inexistant(s)\n")
    sys.exit(3)


################################################################
# construire une liste "candidats" à partir de tous les fichiers Cyclade
candidats = []
for data in files_cyclade:
    with open(data, encoding='utf-8-sig') as f:
        reader = list(csv.reader(f, delimiter=';', quotechar="'"))
        # liste de liste ; reader[r] : chaque ligne ; reader[r][c] : chaque cellule
        session     = reader[0][0]
        etab        = reader[2][0]
        etab_nom    = etab.split('(')[0][:-1]
        etab_uai    = etab.split('(')[1][:-1]
        ''' les données sont structurées ainsi :
        ['Division de classe', 'N° Candidat', 'N° Inscription', 'N° Océan', 'Nom de famille',
        "Nom d'usage", 'Prénom(s)', 'Date de Naissance', 'Division de classe', 'INE',
        'Catégorie Candidat', 'Code Spécialité', 'Spécialité',
        'Etat', 'Enseignements']
        exemple :
        ['TMA', '01216557741', '002 Version 2', ' -', 'DURAND',
        ' -', 'Bryan', '09/03/2005', 'TMA', '081277848GG',
        'SCOLAIRE BACPRO 3 ANS (132)', '31212', 'Métiers de l'accueil',
        'Inscrit', 'Non renseigné'] '''
        reader = reader[9:]     # les données commencent à la ligne 10
        for line in reader:
            # candidat : liste au format
            # ['Nom', 'Prénom', 'Date de Naissance', 'N° Candidat', 'Division', 'Code']
            candidat = [line[4], line[6], line[7], line[1], line[0], line[11]]
            candidats.append(candidat)


################################################################
# la liste 'candidats' ne doit pas être vide
if len(candidats) == 0:
    print(f"❌ Un problème est survenu (pas de candidat trouvé) !\n")
    sys.exit(4)


################################################################
# afficher les infos 'établissement' et 'candidats'
info_etab   =   "Infos établissement trouvées :\n"
info_etab   +=  "session            : " + session + "\n"
info_etab   +=  "Nom établissement  : " + etab_nom + "\n"
info_etab   +=  "UAI établissement  : " + etab_uai
info(f"{len(candidats)} candidats trouvés.")
info(info_etab)


touche()


################################################################
# on dispose désormais d'une liste "candidats" ;
# [ [ 'Nom', 'Prénom', 'Date de Naissance', 'N° Candidat', 'Division', 'Code' ], etc. ]
# chaque élément est lui-même une liste, ie. un candidat
#
# ainsi que des variables "globales" : session ; etab_nom ; etab_uai


################################################################
# extraction des diplômes qui concernent l'établissement
diplomes = []
for candidat in candidats:
    diplome = candidat[5]
    if not (diplome in diplomes):
        diplomes.append(diplome)


################################################################
# la liste 'diplomes' ne doit pas être vide
if len(diplomes) == 0:
    print(f"❌ Un problème est survenu (pas de diplôme trouvé) !\n")
    sys.exit(5)


################################################################
# tous les éléments de la liste 'diplomes' doivent être dans les clés du dictionnaire 'DIPLOMES' (cf. constantes)
for diplome in diplomes:
    if not (diplome in DIPLOMES.keys()):
        print(f"Un diplôme inconnu est trouvé : {diplome}.")
        sys.exit(6)


################################################################
# affichage des diplômes trouvés (code + intitulé)
info_diplomes = str(len(diplomes)) + " diplôme(s) trouvé(s) :\n"
for d in diplomes:
    info_diplomes += d + ' : '
    info_diplomes += DIPLOMES[d] + '\n'
info(info_diplomes[:-1])


touche()


################################################################
# vérification de l'existence des fichiers modèles pour chaque diplôme
for d in diplomes:
    if d + '.xlsx' not in files:
        print(f"❌ Un fichier modèle est manquant : {d + '.xlsx'} !\n")
        sys.exit(7)


################################################################
# vérification de l'existence d'une feuille TEMPLATE_SHEET dans chaque fichier modèle
for d in diplomes:
    classeur = d + ".xlsx"
    wb = openpyxl.load_workbook(classeur, read_only=True, data_only=True)
    if TEMPLATE_SHEET not in wb.sheetnames:
        print(f"❌ Le fichier \"{classeur}\" doit posséder une feuille \"{TEMPLATE_SHEET}\" !\n")
        sys.exit(8)
    wb.close()


################################################################
# création de l'arborescence pour les fichiers individuels des candidats
#
# si le dossier existe, le renommer
if os.path.exists(INDIV_FOLDER + etab_uai):
    t = stamp()
    print(f"⚠️ Le répertoire \"{INDIV_FOLDER + etab_uai}\" existe déjà :\nil a été renommé en \"{INDIV_FOLDER}_old_" + t + "\".\n")
    os.rename(INDIV_FOLDER + etab_uai, INDIV_FOLDER + "_old_" + t)
# créer le dossier candidats_UAI
print(f"🟢 Création du répertoire \"{INDIV_FOLDER + etab_uai}\".\n")
os.mkdir(INDIV_FOLDER + etab_uai)
# créer un sous dossier par diplôme
for diplome in diplomes:
    folderName  =   INDIV_FOLDER + etab_uai
    folderName  +=  "/"
    folderName  +=  diplome + "-"
    folderName  +=  DIPLOMES_COURTS[diplome]
    os.mkdir(folderName)


touche()


################################################################
# création des fichiers individuels des candidats dans l'arborescence

# pour mémoire :
# candidats = [ [ 'Nom', 'Prénom', 'Date de Naissance', 'N° Candidat', 'Division', 'Code' ], etc. ]
# + variables "globales" : session ; etab_nom ; etab_uai
# arbo =    .    /    INDIV_FOLDER + etab_uai    /    diplome + "-" DIPLOMES_COURTS[diplome]
# arbo =    .    /    candidats_0921500F         /    31212-bacpro_MA
# nom+prenom+code+ncandidat.xlsx
info("Traitement : création des fichiers individuels des candidats")
for candidat in candidats:
    ################################################################
    # copie du fichier 'modèle' vers le fichier 'candidat' dans le bon sous-dossier
    folder      =  "./" + INDIV_FOLDER + etab_uai + "/"
    folder      += candidat[5] + "-"
    folder      += DIPLOMES_COURTS[candidat[5]] + "/"
    filename    =  sanitize(candidat[0]) + "+"
    filename    += sanitize(candidat[1]) + "+"
    filename    += candidat[5] + "+"
    filename    += candidat[3] + ".xlsx"
    print("\n" + "-" * 32)
    print(f"Candidat traité : {candidat[0]} {candidat[1]}, né(e) le {candidat[2]}")
    print(f"Diplôme : {DIPLOMES_COURTS[candidat[5]]} (code : {candidat[5]})")
    print(f"Division : {candidat[4]} - Numéro de candidat : {candidat[3]}")
    print(f"Nom du dossier : {folder}")
    print(f"Nom du fichier : {filename}")
    # exemple source :
    #   31212.xlsx
    # exemple destination :
    #   ./0921234A_candidats/31212-bacpro_MA/DURAND+Clara+31212+06916557742.xlsx
    source      = candidat[5] + ".xlsx"
    destination = folder + filename
    # print(f"{source} -> {destination}")
    shutil.copyfile(source, destination)
    time.sleep(TEMPO)  # pour 'terminer' l'écriture du fichier
    #
    ################################################################
    # personnalisation des fichiers candidats (insertion des valeurs)
    # pour mémoire :
    # clés de TEMPLATE_DICT :
    # 'session', 'etab', 'UAI', 'nom', 'prenom', 'daten', 'numcandidat', 'division', 'code'
    # pour mémoire :
    # candidats = [ [ 'Nom', 'Prénom', 'Date de Naissance', 'N° Candidat', 'Division', 'Code' ], etc. ]
    #
    wb = openpyxl.load_workbook(destination, read_only=False)
    sheet = wb[TEMPLATE_SHEET]
    sheet[TEMPLATE_DICT['session']]     = session
    sheet[TEMPLATE_DICT['etab']]        = etab_nom
    sheet[TEMPLATE_DICT['UAI']]         = etab_uai
    sheet[TEMPLATE_DICT['nom']]         = candidat[0]
    sheet[TEMPLATE_DICT['prenom']]      = candidat[1]
    sheet[TEMPLATE_DICT['daten']]       = candidat[2]
    sheet[TEMPLATE_DICT['numcandidat']] = candidat[3]
    sheet[TEMPLATE_DICT['division']]    = candidat[4]
    sheet[TEMPLATE_DICT['code']]        = candidat[5]
    wb.save(destination)
    wb.close()

msg_fin = f"""

🟢 Les fichiers des candidats sont créés :

Dans le dossier "{INDIV_FOLDER}{etab_uai}", un sous-dossier est
préparé par diplôme.
Chacun d'entre eux contient les fichiers individuels des candidats,
avec les informations nominatives mises à jour.

That's all folks!

"""

info(msg_fin)

# fin










'''

*************************************
    REMARQUES
*************************************

👉 : avec un fichier Excel 'propre', cela n'arrive plus !

WARNING en début de traitement :

Élève traité : COLEMAN - Leslie - M2023094837 - CAP_EPC
        Nom du fichier : CAP_EPC+COLEMAN+Leslie+M2023094837.xlsx
/usr/lib/python3/dist-packages/openpyxl/worksheet/_reader.py:300: UserWarning: Data Validation extension is not supported and will be removed
  warn(msg)
/usr/lib/python3/dist-packages/openpyxl/worksheet/_reader.py:300: UserWarning: Unknown extension is not supported and will be removed
  warn(msg)

*************************************

👉 : avec un fichier Excel 'propre', cela n'arrive plus !

taille énorme des fichiers XLSX : 14 Mo !!
Pourquoi ? Lié au pb ci-dessus ?

*************************************

'''




'''
*************************************
    old code :
*************************************

# vérification : la feuille TEMPLATE_SHEET existe dans le fichier TEMPLATE_FILE
wb = openpyxl.load_workbook(TEMPLATE_FILE, read_only=True, data_only=True)
if TEMPLATE_SHEET not in wb.sheetnames:
    print(f"❌ Le fichier \"{TEMPLATE_FILE}\" doit posséder une feuille \"{TEMPLATE_SHEET}\" !\n")
    sys.exit()
wb.close()

# vérification : la feuille LIST_SHEET existe dans le fichier LIST_FILE
wb = openpyxl.load_workbook(LIST_FILE, read_only=True, data_only=True)
if LIST_SHEET not in wb.sheetnames:
    print(f"❌ Le fichier \"{LIST_FILE}\" doit posséder une feuille \"{LIST_SHEET}\" !\n")
    sys.exit()

# récupération des données dans une liste "data"
sheet = wb[LIST_SHEET]
data = []
for l in range(LIST_RANGE[0], LIST_RANGE[2]+1):
    eleve = []
    not_empty = True  # la ligne n'est pas vide (ie : pas de valeur dans la première colonne)
    for c in range(LIST_RANGE[1], LIST_RANGE[3]+1):
        cell = sheet.cell(row=l, column=c).value
        if cell is None:
            not_empty = False
        else:
            eleve.append(cell)
    if not_empty:
        data.append(eleve)
examen = sanitize(sheet.cell(row=1, column=1).value)
wb.close()
# print(data, examen)

for eleve in data:
    [matricule, nom, prenom] = eleve
    print(f"\n\nÉlève traité : {nom} - {prenom} - {matricule} - {examen}")
    # copie du fichier TEMPLATE_FILE dans le répertoire INDIV_FOLDER, nommé examen+nom+prenom.ncandidat.xlsx
    filename  = sanitize(examen) + CHAR_SEP
    filename += sanitize(nom) + CHAR_SEP
    filename += sanitize(prenom) + CHAR_SEP
    filename += matricule + ".xlsx"
    print("\tNom du fichier :", filename)
    shutil.copyfile(TEMPLATE_FILE, INDIV_FOLDER+'/'+filename)
    time.sleep(TEMPO)
    
'''
